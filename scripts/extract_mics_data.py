import json
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT.parent / "outputs" / "mics-rounds-2-6" / "UNICEF_MICS_Rounds_2_to_6_Question_Include.xlsx"
OUTPUT = ROOT / "public" / "data" / "mics-question-include.json"


def clean(value):
    if value is None:
        return ""
    return str(value).strip()


workbook = load_workbook(SOURCE, read_only=True, data_only=True)
sheet = workbook["Question Include"]
headers = [clean(cell.value) for cell in next(sheet.iter_rows())]

rows = []
for values in sheet.iter_rows(min_row=2, values_only=True):
    record = dict(zip(headers, values))
    rows.append(
        {
            "contentTitle": clean(record.get("Content Title")),
            "region": clean(record.get("Region")),
            "survey": clean(record.get("Survey")),
            "round": clean(record.get("MICS Round")),
            "question": clean(record.get("Question")),
            "include": 1 if record.get("Include") == 1 else 0,
        }
    )

payload = {
    "source": SOURCE.name,
    "columns": ["contentTitle", "region", "survey", "round", "question", "include"],
    "rows": rows,
}

OUTPUT.parent.mkdir(parents=True, exist_ok=True)
OUTPUT.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
print(json.dumps({"output": str(OUTPUT), "rows": len(rows), "bytes": OUTPUT.stat().st_size}))
